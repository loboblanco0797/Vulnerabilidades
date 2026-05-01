import pandas as pd
import requests
import logging
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill 
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Configuración de logs profesionales
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class VulnerabilityDashboard:
    def __init__(self):
        self.url = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
        self.df = None
        self.filename = f"Dashboard_SOC_CISA_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Estilos base
        self.font_header = Font(name="Segoe UI", size=20, bold=True, color="FFFFFF")
        self.font_button_title = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    def fetch_data(self):
        """Descarga e inteligencia de datos."""
        try:
            logging.info("Sincronizando feed CISA KEV...")
            response = requests.get(self.url, timeout=15)
            response.raise_for_status()
            self.df = pd.DataFrame(response.json()['vulnerabilities'])
            
            # Sanitización de datos
            for col in self.df.columns:
                self.df[col] = self.df[col].apply(lambda x: ', '.join(map(str, x)) if isinstance(x, list) else x)
            
            # Convertir a datetime para cálculos temporales
            self.df['dateAdded'] = pd.to_datetime(self.df['dateAdded'])
            logging.info(f"Datos cargados: {len(self.df)} registros.")
            return True
        except Exception as e:
            logging.error(f"Falla en obtención de datos: {e}")
            return False

    def setup_dashboard_sheet(self, wb):
        """Diseña la interfaz interactiva con filtros temporales visibles."""
        ws = wb.active
        ws.title = "Dashboard"
        ws.sheet_view.showGridLines = False 
        
        # Paleta SOC
        bg_dark = "0B192C"      
        btn_blue = "1E3A8A"     
        text_white = "FFFFFF"   
        
        # Definir "Reciente" (últimos 30 días)
        fecha_limite = datetime.now() - timedelta(days=30)

        # 1. Fondo oscuro
        for r in range(1, 45):
            for c in range(1, 20):
                ws.cell(row=r, column=c).fill = PatternFill(start_color=bg_dark, end_color=bg_dark, fill_type="solid")

        # 2. Título Principal
        ws.merge_cells("C2:M3")
        title_cell = ws["C2"]
        title_cell.value = "CENTRO DE INTELIGENCIA DE AMENAZAS - CISA KEV"
        title_cell.font = self.font_header
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtítulo de periodo de datos
        ws.merge_cells("C4:M4")
        sub_cell = ws["C4"]
        sub_cell.value = f"Análisis basado en explotación activa confirmada | Reporte generado: {datetime.now().strftime('%d/%m/%Y')}"
        sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="CCCCCC")
        sub_cell.alignment = Alignment(horizontal='center')

        # 3. Cards con Inteligencia de Fechas
        categories_data = [
            ("MICROSOFT", "Vulnerabilidades_Microsoft", "4F81BD", 
             self.df['vendorProject'].str.contains('Microsoft', case=False)),
            ("CISCO", "Vulnerabilidades_Cisco", "3A86FF", 
             self.df['vendorProject'].str.contains('Cisco', case=False)),
            ("APPLE", "Vulnerabilidades_Apple", "8E9AAF", 
             self.df['vendorProject'].str.contains('Apple', case=False)),
            ("CLOUD", "Vulnerabilidades_Cloud", "00B4D8", 
             self.df['vendorProject'].str.contains('AWS|Cloud|Google|Azure', case=False)),
            ("OPEN SOURCE", "Vulnerabilidades_OSS", "FB8500", 
             self.df['shortDescription'].str.contains('Open Source|Librer|OpenSSL', case=False)),
            ("OTROS", "Vulnerabilidades_Otros", "8338EC", 
             ~self.df['vendorProject'].str.contains('Microsoft|Cisco|Apple|AWS|Cloud', case=False))
        ]
        
        start_row, start_col = 7, 3
        
        for index, (title, target, accent_color, mask) in enumerate(categories_data):
            r = start_row + (index // 3) * 8
            c = start_col + (index % 3) * 4
            
            # Cálculos dinámicos
            df_cat = self.df[mask]
            total = len(df_cat)
            recientes = len(df_cat[df_cat['dateAdded'] >= fecha_limite])
            
            # Texto del Botón
            main_cell = ws.cell(row=r, column=c)
            main_cell.value = (
                f"{title}\n"
                f"──────────────────\n"
                f"Histórico Total: {total}\n"
                f"Nuevas (Últimos 30d): {recientes}"
            )
            main_cell.font = self.font_button_title
            main_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            main_cell.hyperlink = f"#{target}!A1"
            
            # Estilo del Card
            card_border = Border(left=Side(style='thick', color=accent_color), 
                                 right=Side(style='thick', color=accent_color), 
                                 top=Side(style='thick', color=accent_color), 
                                 bottom=Side(style='thick', color=accent_color))
            
            for row_btn in range(r, r + 6):
                for col_btn in range(c, c + 3):
                    cell = ws.cell(row=row_btn, column=col_btn)
                    cell.fill = PatternFill(start_color=btn_blue, end_color=btn_blue, fill_type="solid")
                    cell.border = card_border
            
            ws.merge_cells(f"{get_column_letter(c)}{r}:{get_column_letter(c+2)}{r+5}")
        
        for col_num in range(3, 14):
            ws.column_dimensions[get_column_letter(col_num)].width = 25

    def add_data_sheets(self, wb):
        """Crea hojas de datos con formato de tabla."""
        categories = [
            ("Vulnerabilidades_Microsoft", self.df['vendorProject'].str.contains('Microsoft', case=False)),
            ("Vulnerabilidades_Cisco", self.df['vendorProject'].str.contains('Cisco', case=False)),
            ("Vulnerabilidades_Apple", self.df['vendorProject'].str.contains('Apple', case=False)),
            ("Vulnerabilidades_Cloud", self.df['vendorProject'].str.contains('AWS|Cloud|Google|Azure', case=False)),
            ("Vulnerabilidades_OSS", self.df['shortDescription'].str.contains('Open Source|Librer|OpenSSL', case=False)),
            ("Vulnerabilidades_Otros", ~self.df['vendorProject'].str.contains('Microsoft|Cisco|Apple|AWS|Cloud', case=False))
        ]
        
        for sheet_name, mask in categories:
            df_filtered = self.df[mask].copy()
            if not df_filtered.empty:
                # Volver a formato fecha corta para Excel
                df_filtered['dateAdded'] = df_filtered['dateAdded'].dt.date
                ws = wb.create_sheet(title=sheet_name)
                ws.append(list(df_filtered.columns))
                for row in df_filtered.values:
                    ws.append(list(row))
                
                full_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                tab = Table(displayName=f"Tabla{sheet_name.replace('_','')}", ref=full_range)
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tab)
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 35

    def create_dashboard(self):
        """Genera y abre el Dashboard."""
        wb = Workbook()
        self.setup_dashboard_sheet(wb)
        self.add_data_sheets(wb)
        try:
            wb.save(self.filename)
            logging.info("¡Dashboard con inteligencia temporal generado!")
            os.startfile(self.filename)
        except Exception as e:
            logging.error(f"Error al guardar: {e}")

if __name__ == "__main__":
    app = VulnerabilityDashboard()
    if app.fetch_data():
        app.create_dashboard()